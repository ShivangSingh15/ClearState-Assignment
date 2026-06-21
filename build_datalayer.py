"""
build_datalayer.py
==================
Assembles Roche_Dashboard_DataLayer.xlsx from FACT_long.csv + the raw RoFIS
workbook. One-time build step, run after normalize.py.

Run:   python3 normalize.py        # -> FACT_long.csv
       python3 build_datalayer.py  # -> Roche_Dashboard_DataLayer.xlsx

Design note — columns are NOT hardcoded
---------------------------------------
02_FACT is written by streaming FACT_long.csv row-for-row: the header and every
column come straight from the CSV, so if normalize.py adds/renames a column the
sheet follows automatically. 03_DIM_Entity is derived from the same CSV via
DictReader. The only literal column lists here belong to the two CONFIG tabs,
whose columns ARE the schema (key/value/description and the company settings).

Tabs produced (12)
  00_CONFIG      global settings (FY fallback, defaults, source, last_refresh)
  01_COMPANY     one row per company (company-specific settings; overrides 00_CONFIG)
  02_FACT        normalised analytical rows, verbatim from FACT_long.csv
  03_DIM_Entity  unique entity hierarchy derived from FACT (company-prefixed)
  RAW_*          8 verbatim source sheets for traceability
"""

import openpyxl, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = 'roche-group-financial-data.xlsx'
FACT_CSV = 'FACT_long.csv'
OUT = 'Roche_Dashboard_DataLayer.xlsx'

try:
    src = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
except FileNotFoundError:
    src = None   # RAW tabs skipped if the raw workbook isn't alongside this script
wb = Workbook(); wb.remove(wb.active)

# ── shared styles ────────────────────────────────────────────────────────────
HDR     = Font(bold=True, color='FFFFFF', name='Arial', size=10)
HDRFILL = PatternFill('solid', start_color='1F4E5F')
BLUE    = Font(color='0000FF', name='Arial', size=10)
NORM    = Font(name='Arial', size=10)
ITALIC  = Font(italic=True, color='888888', name='Arial', size=10)
ROCHEFILL = PatternFill('solid', start_color='EBF3FA')
thin = Side(style='thin', color='D0D0D0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, ncol):
    """Style row 1 as a header and freeze it. ncol is read from the data, not fixed."""
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = HDR
        cell.fill = HDRFILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    ws.freeze_panes = 'A2'


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def autofilter(ws, ncol):
    """Filter across all columns present — range follows the data width."""
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}1"


# ════════════════════════════════════════════════════════════════════════════
# 00_CONFIG  — global settings (fallbacks + defaults + source + policies)
# ════════════════════════════════════════════════════════════════════════════
cfg = wb.create_sheet('00_CONFIG')
cfg.append(['key', 'value', 'description'])
config_rows = [
 ['fiscal_year_start_month','1','Global fallback — each company sets its own in 01_COMPANY'],
 ['fiscal_year_label_mode','start_year','How to label a fiscal year that spans two calendar years: start_year or end_year'],
 ['default_basis','CHF','Default measurement basis: CHF (reported) or CER (constant exchange rate)'],
 ['default_period_type','FullYear','Default period view: Quarter | YTD | FullYear'],
 ['default_entity','Group','Default entity selected on load'],
 ['default_year','2025','Default reporting year on load'],
 ['currency_unit','CHF m','Display unit for all values'],
 ['company','Roche','Global default — overridden per company in 01_COMPANY tab'],
 ['source_name','Roche Finance Information Tool (RoFIS)','Public source system'],
 ['source_url','https://www.roche.com/investors/rofis','Source URL'],
 ['source_file','roche-group-financial-data.xlsx','Downloaded workbook file name'],
 ['source_release','Q1 2026 RoFIS release','Reporting period of the downloaded workbook'],
 ['retrieval_date','2026-06-21','Date data was retrieved from the source'],
 ['last_refresh','2026-06-21 00:00','Set by the refresh routine when FACT is rebuilt'],
 ['cer_methodology','Roche restates prior-period sales at the comparison period average FX. CER values and CER growth are taken AS PUBLISHED; never derived in this tool.','Methodology note shown in the CER tooltip'],
 ['incomplete_year_policy','A FullYear view is flagged INCOMPLETE if fewer than 4 standalone quarters are present for that year.','Drives the incomplete-data badge'],
]
for r in config_rows:
    cfg.append(r)
style_header(cfg, 3)
set_widths(cfg, [26, 46, 70])
for r in range(2, cfg.max_row + 1):
    cfg.cell(r, 2).font = BLUE
    for c in range(1, 4):
        cfg.cell(r, c).border = BORDER
        cfg.cell(r, c).alignment = Alignment(vertical='top', wrap_text=(c == 3))

# ════════════════════════════════════════════════════════════════════════════
# 01_COMPANY — one row per company; getConfig_() overrides 00_CONFIG with this
# ════════════════════════════════════════════════════════════════════════════
co = wb.create_sheet('01_COMPANY')
company_headers = ['company','fiscal_year_start_month','fiscal_year_label_mode',
                   'currency_unit','default_entity','default_year','default_basis',
                   'cer_methodology','source_name','source_url','source_file','source_release']
co.append(company_headers)
co.append(['Roche',1,'start_year','CHF m','Group',2025,'CHF',
           'Roche restates prior-period sales at the comparison period average FX. '
           'CER values and CER growth are taken AS PUBLISHED; never derived in this tool.',
           'Roche Finance Information Tool (RoFIS)',
           'https://www.roche.com/investors/rofis',
           'roche-group-financial-data.xlsx','Q1 2026 RoFIS release'])
# Placeholder row — leading '#' makes getCompanyList_() skip it. Replace to onboard.
co.append(['# Example: Novartis',1,'start_year','USD m','Group',2025,'CHF',
           'Novartis CER methodology note here.','Novartis Investor Relations',
           'https://www.novartis.com/investors','novartis-financial-data.xlsx','Q1 2026 release'])
style_header(co, len(company_headers))
set_widths(co, [18,24,22,14,16,14,14,60,36,44,36,22])
for c in range(1, len(company_headers) + 1):
    co.cell(2, c).fill = ROCHEFILL            # Roche row highlighted
    co.cell(3, c).font = ITALIC               # example row greyed/italic

# ════════════════════════════════════════════════════════════════════════════
# 02_FACT — streamed verbatim from the CSV (columns + values come from the file)
# ════════════════════════════════════════════════════════════════════════════
fact = wb.create_sheet('02_FACT')
rows = list(csv.reader(open(FACT_CSV)))
for r in rows:
    fact.append(r)                            # values stay as text, exactly as in CSV
ncol = len(rows[0])
style_header(fact, ncol)
# widths align to the known column order in the CSV header; harmless if it grows
set_widths(fact, [10,16,26,16,22,16,8,11,14,9,12,12,12,11,17,17,12,11,28,30,26][:ncol])
autofilter(fact, ncol)

# ════════════════════════════════════════════════════════════════════════════
# 03_DIM_Entity — unique entities derived from FACT (company kept as first column)
# ════════════════════════════════════════════════════════════════════════════
dim = wb.create_sheet('03_DIM_Entity')
dim_cols = ['company','entity','entity_level','division','parent_entity','drilldown_dim']
dim.append(dim_cols)
seen = set()
for r in csv.DictReader(open(FACT_CSV)):
    k = tuple(r[c] for c in dim_cols)         # unique key over the dimension columns
    if k in seen:
        continue
    seen.add(k)
    dim.append(list(k))
style_header(dim, len(dim_cols))
set_widths(dim, [14,18,16,24,18,13])
autofilter(dim, len(dim_cols))

# ════════════════════════════════════════════════════════════════════════════
# RAW_*  — verbatim source sheets (traceability for every source_ref in FACT)
# ════════════════════════════════════════════════════════════════════════════
raw_map = {
 'Group Sales CHF':'RAW_GroupSales_CHF',          'Group Sales CER':'RAW_GroupSales_CER',
 'P Sales Global CHF':'RAW_PharmaProducts_CHF',   'P Sales Global CER':'RAW_PharmaProducts_CER',
 'P Therapeutic areas CHF':'RAW_PharmaTA_CHF',    'P Therapeutic areas CER':'RAW_PharmaTA_CER',
 'Dia Sales CHF':'RAW_DiaSales_CHF',              'Dia Sales CER':'RAW_DiaSales_CER',
}
if src is not None:
    for srcname, tab in raw_map.items():
        ws_s = src[srcname]
        ws = wb.create_sheet(tab)
        for r in range(1, ws_s.max_row + 1):
            ws.append([ws_s.cell(r, c).value for c in range(1, ws_s.max_column + 1)])
        ws.insert_rows(1)
        ws.cell(1, 1).value = (f'RAW provenance — copied verbatim from {SRC.split("/")[-1]} '
                               f':: sheet "{srcname}". Do not edit; FACT is derived from this.')
        ws.cell(1, 1).font = Font(italic=True, color='888888', size=9, name='Arial')
else:
    print('NOTE: raw workbook not found — RAW_* tabs skipped (run beside '
          'roche-group-financial-data.xlsx to include them).')

wb.save(OUT)
print('saved. tabs:', wb.sheetnames)
print('FACT data rows:', len(rows) - 1, '| DIM rows:', dim.max_row - 1)

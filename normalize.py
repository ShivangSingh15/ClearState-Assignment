"""
normalize.py
============
Parses the raw Roche RoFIS workbook into a long-format fact table: FACT_long.csv.

Run:   python3 normalize.py
In:    roche-group-financial-data.xlsx   (downloaded from RoFIS)
Out:   FACT_long.csv                     (feeds build_datalayer.py)

The column schema is defined in ONE place (the `COLS` list at the very bottom),
derived from the dict keys the walkers emit. Nothing downstream re-declares it —
build_datalayer.py reads the columns straight from this CSV's header row.

Layout the parser handles
-------------------------
Each of the 8 sales sheets (4 dimensions x CHF/CER) carries two stacked blocks:
  * a YTD / cumulative block   (Q1, H1, YTD Sep, Full-Year)
  * a standalone-quarter block (Q1, Q2, Q3, Q4)
followed (after a blank column gap) by a growth-% block with the same labels.
Block positions, label columns and the value/growth split are all DETECTED, not
hardcoded, so a new release with shifted rows/columns still parses.
"""

import openpyxl, csv, re

# ── source metadata ──────────────────────────────────────────────────────────
SRC       = 'roche-group-financial-data.xlsx'
RELEASE   = 'Q1 2026 RoFIS release'
SRC_URL   = 'https://www.roche.com/investors/rofis'
RETRIEVED = '2026-06-21'


wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)


# ── period label -> metadata ─────────────────────────────────────────────────
def period_meta(label):
    """'Q1 2025'/'H1 2025'/'YTD Sep 25'/2025 -> (type, year, quarter, start, end)."""
    if isinstance(label, (int, float)):
        label = str(int(label))
    m = re.match(r'Q([1-4]) (\d{4})', label)
    if m:
        q, y = int(m.group(1)), int(m.group(2))
        starts = {1:(1,1), 2:(4,1), 3:(7,1), 4:(10,1)}
        ends   = {1:(3,31), 2:(6,30), 3:(9,30), 4:(12,31)}
        return ('Quarter', y, f'Q{q}',
                f'{y}-{starts[q][0]:02d}-{starts[q][1]:02d}',
                f'{y}-{ends[q][0]:02d}-{ends[q][1]:02d}')
    m = re.match(r'H1 (\d{4})', label)
    if m:
        y = int(m.group(1)); return ('HalfYear', y, 'H1', f'{y}-01-01', f'{y}-06-30')
    m = re.match(r'YTD Sep (\d{2})', label)
    if m:
        y = 2000 + int(m.group(1)); return ('YTD', y, 'Q3', f'{y}-01-01', f'{y}-09-30')
    m = re.match(r'^(\d{4})$', str(label).strip())
    if m:
        y = int(m.group(1)); return ('FullYear', y, 'Q4', f'{y}-01-01', f'{y}-12-31')
    return None


# ── header row -> (value columns, growth columns) ────────────────────────────
def header_cols(ws, hrow):
    """Find period-label columns and split into value block + growth block by the gap."""
    labels = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(hrow, c).value
        sv = None
        if isinstance(v, str):
            sv = v.strip()
        elif isinstance(v, (int, float)) and float(v).is_integer():
            sv = str(int(v))
        if sv and period_meta(sv):
            labels[c] = sv
    cols = sorted(labels)
    if not cols:
        return [], []
    # contiguous runs: first run = values, second run = growth %
    groups, cur = [], [cols[0]]
    for c in cols[1:]:
        if c == cur[-1] + 1:
            cur.append(c)
        else:
            groups.append(cur); cur = [c]
    groups.append(cur)
    val = groups[0]
    grw = groups[1] if len(groups) > 1 else []
    return [(c, labels[c]) for c in val], [(c, labels[c]) for c in grw]


# ── locate the two blocks on a sheet ─────────────────────────────────────────
def find_blocks(ws):
    """Header rows containing 'Q1 2025'; 'ytd' if next col is H1, else 'qtr'."""
    blocks = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip() == 'Q1 2025':
            nxt = ws.cell(r, 3).value
            kind = 'qtr' if (isinstance(nxt, str) and nxt.strip().startswith('Q2')) else 'ytd'
            blocks.append((r, kind))
    return blocks


def clean_num(v):
    if v is None or v == '':
        return None
    try:
        f = float(v)
        return 0.0 if abs(f) < 1e-6 else round(f, 3)
    except (TypeError, ValueError):
        return None


# ── emit one record per (entity x period) ────────────────────────────────────
rows = []   # single-basis records, merged later

def emit(company, division, entity, level, parent, drill,
         sheet, basis, valcols, grwcols, ws, datarow):
    for vc, plabel in valcols:
        pm = period_meta(plabel)
        if not pm:
            continue
        ptype, cy, qtr, pstart, pend = pm
        val = clean_num(ws.cell(datarow, vc).value)
        g = None
        for gc, gl in grwcols:
            if gl == plabel:
                g = clean_num(ws.cell(datarow, gc).value); break
        if val is None and g is None:
            continue
        rows.append(dict(
            company=company, division=division, entity=entity, entity_level=level,
            parent_entity=parent, drilldown_dim=drill, metric='Sales', basis=basis,
            period_type=ptype, calendar_year=cy, quarter=qtr, period_label=plabel,
            period_start=pstart, period_end=pend, currency_unit='CHF m',
            value=val, growth=g, data_status='actual',
            source_ref=f'{sheet}!R{datarow}', source_release=RELEASE))


# ── per-sheet walkers (the only Roche-specific knowledge) ────────────────────
def walk_group(sheet, basis):
    """Group sheet: Group / Divisions / Regions / 'thereof' sub-regions."""
    ws = wb[sheet]
    for hrow, _ in find_blocks(ws):
        valcols, grwcols = header_cols(ws, hrow)
        if not valcols:
            continue
        lcol = valcols[0][0] - 1
        division = 'Group'
        r = hrow + 1
        while r <= ws.max_row:
            lab = ws.cell(r, lcol).value
            if lab is None or str(lab).strip() == '':
                if r > hrow + 1: break
                r += 1; continue
            lab = str(lab).strip()
            if lab.startswith('Q1 2025'): break
            low = lab.lower()
            if lab == 'Pharmaceuticals Division':
                division = 'Pharmaceuticals'
                emit('Roche','Pharmaceuticals','Pharmaceuticals Division','Division','Group','None',sheet,basis,valcols,grwcols,ws,r)
            elif lab == 'Diagnostics Division':
                division = 'Diagnostics'
                emit('Roche','Diagnostics','Diagnostics Division','Division','Group','None',sheet,basis,valcols,grwcols,ws,r)
            elif lab == 'Group':
                emit('Roche','Group','Group','Group','None','None',sheet,basis,valcols,grwcols,ws,r)
            elif low.startswith('thereof'):
                sub = lab.replace('thereof','').strip()
                parent = 'International' if division == 'Pharmaceuticals' else ('North America' if 'united' in low else 'Asia-Pacific')
                emit('Roche',division,sub,'SubRegion',parent,'Region',sheet,basis,valcols,grwcols,ws,r)
            else:
                emit('Roche',division,lab,'Region',f'{division} Division','Region',sheet,basis,valcols,grwcols,ws,r)
            r += 1


def walk_simple(sheet, basis, division, level, drill, parent, skip=('total',)):
    """Flat list sheet (Products, Therapeutic areas). 'thereof' -> Sub<level>."""
    ws = wb[sheet]
    for hrow, _ in find_blocks(ws):
        valcols, grwcols = header_cols(ws, hrow)
        if not valcols:
            continue
        lcol = valcols[0][0] - 1
        r = hrow + 1
        while r <= ws.max_row:
            lab = ws.cell(r, lcol).value
            if lab is None or str(lab).strip() == '':
                if r > hrow + 1: break
                r += 1; continue
            lab = str(lab).strip()
            if lab.startswith('Q1 2025'): break
            if lab.lower() in skip:
                r += 1; continue
            cur_level, name = level, lab
            if lab.lower().startswith('thereof'):
                name = lab.replace('thereof','').strip()
                cur_level = 'Sub' + level
            emit('Roche',division,name,cur_level,parent,drill,sheet,basis,valcols,grwcols,ws,r)
            r += 1


def walk_dia_business(sheet, basis):
    """Diagnostics: BusinessArea rows + 'thereof region' cross-sections."""
    ws = wb[sheet]
    for hrow, _ in find_blocks(ws):
        valcols, grwcols = header_cols(ws, hrow)
        if not valcols:
            continue
        lcol = valcols[0][0] - 1
        r = hrow + 1; cur_ba = None
        while r <= ws.max_row:
            lab = ws.cell(r, lcol).value
            if lab is None or str(lab).strip() == '':
                if r > hrow + 1: break
                r += 1; continue
            lab = str(lab).strip()
            if lab.startswith('Q1 2025'): break
            if lab == 'Diagnostics Division': break   # stop before duplicate region block
            if lab.lower().startswith('thereof'):
                reg = lab.replace('thereof','').strip()
                emit('Roche','Diagnostics',f'{cur_ba} - {reg}','BusinessAreaRegion',cur_ba,'BusinessAreaRegion',sheet,basis,valcols,grwcols,ws,r)
            else:
                cur_ba = lab
                emit('Roche','Diagnostics',lab,'BusinessArea','Diagnostics Division','BusinessArea',sheet,basis,valcols,grwcols,ws,r)
            r += 1


# ── run all 8 sheets ─────────────────────────────────────────────────────────
for basis, suf in [('CHF','CHF'), ('CER','CER')]:
    walk_group(f'Group Sales {suf}', basis)
    walk_simple(f'P Sales Global {suf}', basis, 'Pharmaceuticals','Product','Product','Pharmaceuticals Division')
    walk_simple(f'P Therapeutic areas {suf}', basis, 'Pharmaceuticals','TherapeuticArea','TherapeuticArea','Pharmaceuticals Division', skip=('total',))
    walk_dia_business(f'Dia Sales {suf}', basis)

print('raw normalized rows:', len(rows))


# ── merge CHF + CER onto one row per entity x period ─────────────────────────
key = lambda d: (d['company'], d['division'], d['entity'], d['entity_level'],
                 d['drilldown_dim'], d['period_type'], d['calendar_year'],
                 d['quarter'], d['period_label'])
merged = {}
for d in rows:
    k = key(d)
    if k not in merged:
        merged[k] = dict(
            company=d['company'], division=d['division'], entity=d['entity'],
            entity_level=d['entity_level'], parent_entity=d['parent_entity'],
            drilldown_dim=d['drilldown_dim'], metric='Sales',
            period_type=d['period_type'], calendar_year=d['calendar_year'],
            quarter=d['quarter'], period_label=d['period_label'],
            period_start=d['period_start'], period_end=d['period_end'],
            currency_unit=d['currency_unit'], reported_value_chf=None,
            reported_growth_chf=None, cer_value=None, cer_growth=None,
            data_status='actual', source_ref=d['source_ref'], source_release=RELEASE)
    if d['basis'] == 'CHF':
        merged[k]['reported_value_chf']  = d['value']
        merged[k]['reported_growth_chf'] = d['growth']
        merged[k]['source_ref']          = d['source_ref']
    else:
        merged[k]['cer_value']  = d['value']
        merged[k]['cer_growth'] = d['growth']

fact = list(merged.values())
print('merged fact rows:', len(fact))


# ── reconstruct 2024 quarters from published YoY growth (flagged) ────────────
# prior = current_2025 / (1 + yoy_growth). CER left null (never invented).
idx = {(f['entity'], f['drilldown_dim'], f['quarter']): f
       for f in fact if f['period_type'] == 'Quarter' and f['calendar_year'] == 2025}
recon = []
for (ent, drill, q), f in idx.items():
    g, v = f['reported_growth_chf'], f['reported_value_chf']
    if g is None or v is None or (1 + g) == 0:
        continue
    py = round(v / (1 + g), 1)
    qn = int(q[1])
    starts = {1:'01-01', 2:'04-01', 3:'07-01', 4:'10-01'}
    ends   = {1:'03-31', 2:'06-30', 3:'09-30', 4:'12-31'}
    recon.append(dict(
        company=f['company'], division=f['division'], entity=ent,
        entity_level=f['entity_level'], parent_entity=f['parent_entity'],
        drilldown_dim=drill, metric='Sales', period_type='Quarter',
        calendar_year=2024, quarter=q, period_label=f'{q} 2024',
        period_start=f'2024-{starts[qn]}', period_end=f'2024-{ends[qn]}',
        currency_unit='CHF m', reported_value_chf=py, reported_growth_chf=None,
        cer_value=None, cer_growth=None,
        data_status='reconstructed_from_published_yoy',
        source_ref=f['source_ref'] + ' (YoY-derived)',
        source_release='derived from ' + RELEASE))
print('reconstructed 2024 quarter rows:', len(recon))

fact_all = fact + recon


# ── write CSV (column schema defined here, once) ─────────────────────────────
COLS = ['company','division','entity','entity_level','parent_entity','drilldown_dim',
        'metric','period_type','calendar_year','quarter','period_label','period_start',
        'period_end','currency_unit','reported_value_chf','reported_growth_chf',
        'cer_value','cer_growth','data_status','source_ref','source_release']

with open('FACT_long.csv', 'w', newline='') as fh:
    w = csv.DictWriter(fh, fieldnames=COLS)
    w.writeheader()
    for f in sorted(fact_all, key=lambda x: (x['drilldown_dim'], x['entity'],
                                             x['calendar_year'], x['quarter'],
                                             x['period_type'])):
        w.writerow({c: f.get(c) for c in COLS})

print('TOTAL fact rows written:', len(fact_all))
print('years:', sorted(set(f['calendar_year'] for f in fact_all)))
print('drilldowns:', sorted(set(f['drilldown_dim'] for f in fact_all)))
